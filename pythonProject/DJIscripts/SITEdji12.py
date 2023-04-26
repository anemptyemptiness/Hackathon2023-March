def find_bpla_in_dji():
    import requests
    import re
    from bs4 import BeautifulSoup
    from urllib.parse import urlparse
    # import json
    # import os
    import openpyxl
    # import sys
    #
    # try:
    #     workbook = openpyxl.load_workbook('БПЛА_parse.xlsx')
    #     workbook.save('БПЛА_parse.xlsx')
    # except FileNotFoundError:
    #     print()
    # except Exception as e:
    #     print(f"Ошибка при открытии файла: {e}")
    #     sys.exit()



    json_arr = []

    url = "https://www.dji.com/nl/products/camera-drones#mavic-series"
    response = requests.get(url)
    if response.status_code == 200:
        # Вызываем HTML каталога БПЛА и др
        soup = BeautifulSoup(response.content, "html.parser")

        # Получаем все виды серий дронов кроме RC, так как это пульты, а не БПЛА
        sidenav_menu = soup.find("ul", {"class": "sidenav-menu"}).find_all("span")
        name_el_sidenav_menu = [sm.text.replace("Серия ", "") for sm in sidenav_menu if
                                sm.text.replace("Серия ", "") != "RC"]

        # работаем с частью HTML отвечающей за продуктамы из каталога
        soup2 = soup.find_all("div", {"class": "product-content"})
        for drone in soup2:
            # Получаем название устройства
            drones_name = drone.find("h3", {"class": "title"}).text.strip()

            # если есть русские буквы в названии это не дрон:
            if bool(re.search('[а-яА-Я]', drones_name)): continue
            # Была найдена закономерность, если в названии устройства нет названия серии, то устройство не является БПЛА
            # Таким образом оставляем только дроны:
            for substr in name_el_sidenav_menu:
                if substr in drones_name:
                    # Находим ссылку на сайт с информацией о БПЛА
                    drones_ref = drone.find("a", {"class": "dui-learn-more ga-data"})["href"]
                    link_img_drone = drone.find_previous_sibling("a",class_="product-img ga-data")
                    link_img_drone = link_img_drone.find("figure").get('data-layzr')
                    # Преобразуем ссылку что бы вела на характеристики БПЛА:
                    parsed_url = urlparse(drones_ref)  # Разбиваем URL на составляющие
                    new_drones_ref = parsed_url.scheme + "://" + parsed_url.netloc + parsed_url.path + "/specs"  # Заменяем часть строки после вопросительного знака

                    print("\n" + drones_name)
                    print(new_drones_ref)

                    match = re.search(r'\.(\d|[a-zA-Z])+$', link_img_drone).group()
                    link_img_drone = link_img_drone.replace(match, "")
                    print(link_img_drone)

                    # # можно использвать библиотеку from fuzzywuzzy import process для поиска строк немного отличающихся
                    # options = ['apple', 'banana', 'cherry', 'durian']
                    # # ищем похожую строку с опечаткой
                    # search_term = 'appel'
                    # # настраиваем минимальный порог сходства в 70%
                    # threshold = 70
                    # # ищем наиболее похожую строку
                    # best_match, score = process.extractOne(search_term, options)
                    # # проверяем, превышает ли оценка сходства порог
                    # if score >= threshold:
                    #     print(f"Найдено совпадение: {best_match}")
                    # else:
                    #     print("Совпадений не найдено")

                    def find_vallue_in_table(drone_soup, *iskom_strr):
                        for iskom_str in iskom_strr:
                            # найдем ячейку, в которой указано значение iskom_str
                            cell = drone_soup.find('th', string=iskom_str)
                            if cell is not None:
                                # получим значение из следующей за ней ячейки
                                value = cell.find_next_sibling('td').text
                                # print(iskom_str+": " + value)
                                return value
                            # находим элемент с тегом h4 и текстом "Макс. скорость снижения"
                            h4_elem = drone_soup.find('h4', string=iskom_str)
                            if h4_elem is not None:
                                # находим следующий элемент после h4, который содержит значения скоростей
                                speed_elem = h4_elem.parent.find_next_sibling('li')
                                if speed_elem is not None:
                                    speed_elem = speed_elem.find('div')
                                    if speed_elem is not None:
                                        # print(iskom_str+": "+speed_elem.get_text())
                                        return speed_elem.get_text()
                        # print("Не указано: "+iskom_strr[0])
                        return None


                    dron_response = requests.get(new_drones_ref)
                    if response.status_code == 200:
                        drone_soup = BeautifulSoup(dron_response.content, "html.parser")

                        max_speed = find_vallue_in_table(drone_soup, 'Макс. скорость',
                                                     'Макс. скорость (на уровне моря в штиль)', 'Max Speed',
                                                     'Max Flight Velocity', 'Max Flight Speed')
                        max_speed_nabora = find_vallue_in_table(drone_soup, 'Макс. скорость набора высоты',
                                                     'Max Ascent Speed', 'Max Ascent / Descent Speed')
                        max_speed_snizh = find_vallue_in_table(drone_soup, 'Макс. скорость снижения',
                                                     'Максимальная скорость снижения',
                                                     'Макс. скорость вертикального снижения', 'Max Descent Speed',
                                                     'Max Ascent / Descent Speed')
                        max_distanse = find_vallue_in_table(drone_soup, 'Макс. дальность полета',
                                                     'Макс. расстояние полета (в штиль)', 'Макс. расстояние полета')
                        max_hight = find_vallue_in_table(drone_soup, 'Макс. высота полета над уровнем моря',
                                                     'Макс.высота полета над уровнем моря',
                                                     'Максимальная высота взлета над уровнем моря',
                                                     'Максимальная высота взлета',
                                                     'Max Service Ceiling Above Sea Level')
                        energy_eat = find_vallue_in_table(drone_soup, 'Энергопотребление', 'Энергия', 'Энергетика',
                                                     'Energy')  # не нашла только Энергия в аккумуляторе
                        massa = find_vallue_in_table(drone_soup, 'Масса (с аккумулятором и пропеллерами)', 'Масса',
                                                     'Взлетная масса', 'Takeoff Weight',
                                                     'Weight (Battery & Propellers included)',
                                                     'Weight (Battery & Propellers Included)',
                                                     'Take-off Weight')  # масса полезной нагрузки?
                        max_time_fly = find_vallue_in_table(drone_soup, 'Макс. время полета', 'Макс. время полета	',
                                                     'Макс. время полета (в штиль)', 'Max Flight Time',
                                                     'Flight Time')
                        count_vint = find_vallue_in_table(drone_soup,
                                                     'Число винтов')  # не нашла на этом сайте (но по виду у всех 4)

                        json_arr.append({
                            'Название': drones_name,
                            'Ссылка': new_drones_ref,
                            'Картинка': link_img_drone,
                            'Максимальная скорость': max_speed,
                            'Скорость набора': max_speed_nabora,
                            'Скорость снижения': max_speed_snizh,
                            'Максимальная дальность полета': max_distanse,
                            'Максимальная высота полета': max_hight,
                            'Энергопотребление': energy_eat,
                            'Масса полезной погрузки': massa,
                            'Продолжительность полета': max_time_fly,
                            'Число винтов': count_vint

                        })

                    else:
                        print('Ошибка при выполнении запроса')

                    break

            # dron_response = requests.get(new_drones_ref)
            # if response.status_code == 200:
            #     drone_soup = BeautifulSoup(dron_response.content, "html.parser")
            #
            #     propellers_elem = drone_soup.find('span', string='Aantal propellers')
            #     if propellers_elem:
            #         propellers = propellers_elem.find_next('span').text.strip()
            #     else:
            #         propellers = "Unknown"
            #
            #     hovering_div = drone_soup.find('div', class_='product-detail--table')
            #     if hovering_div is not None:
            #         hovering = 'Hovering' in hovering_div.text
            #     else:
            #         hovering = False
            #
            #
            #     print(f'Количество пропеллеров: {propellers}')
            #     print(f'Наличие функции ховеринг: {"Да" if hovering else "Нет"}')
            # else:
            #     print('Ошибка при выполнении запроса')


    else:
        print('Ошибка при выполнении запроса')
    return json_arr

    # def get_name_spase_table(ws):
    #     ws.append(['Название',
    #                'Ссылка',
    #                'Картинка',
    #                'Максимальная скорость',
    #                'Скорость набора',
    #                'Скорость снижения',
    #                'Максимальная дальность полета',
    #                'Максимальная высота полета',
    #                'Энергопотребление',
    #                'Масса полезной погрузки',
    #                'Продолжительность полета',
    #                'Число винтов'])
    #     # first_row_range = ws[1:1]
    #     # # Проходим по каждой ячейке и устанавливаем атрибут protection, который запретит редактирование
    #     # for cell in first_row_range[0]:
    #     #     cell.protection = cell.protection.copy(sort=False)
    # def get_data(ws,json_arr):
    #     for item in json_arr:
    #         ws.append([item['Название'],
    #                    item['Ссылка'],
    #                    item['Картинка'],
    #                    item['Максимальная скорость'],
    #                    item['Скорость набора'],
    #                    item['Скорость снижения'],
    #                    item['Максимальная дальность полета'],
    #                    item['Максимальная высота полета'],
    #                    item['Энергопотребление'],
    #                    item['Масса полезной погрузки'],
    #                    item['Продолжительность полета'],
    #                    item['Число винтов']])
    #
    # try:
    #     workbook = openpyxl.load_workbook('БПЛА_parse.xlsx')
    #     print("Файл успешно открыт")
    #     worksheet = workbook.active
    #     try:
    #         worksheet = workbook['Drones']
    #     except Exception as e:
    #         worksheet = workbook.create_sheet('Drones')
    #     if worksheet.max_row == 1:
    #         get_name_spase_table(worksheet)
    #         get_data(worksheet,json_arr)
    #     else:
    #         get_data(worksheet,json_arr)
    #     # создать список строк, начиная со второй строки (исключая заголовок)
    #     rows = list(worksheet.iter_rows(min_row=2, values_only=True))
    #     # создать список уникальных строк
    #     unique_rows = list(set(map(tuple, rows)))
    #     # очистить содержимое листа
    #     worksheet.delete_rows(2, worksheet.max_row)
    #     # заполнить лист уникальными строками
    #     for row in unique_rows:
    #         worksheet.append(row)
    #     workbook.save('БПЛА_parse.xlsx')
    # except FileNotFoundError:
    #     print("Файл не найден")
    #     wb = openpyxl.Workbook()
    #     ws = wb.create_sheet('Drones')
    #     get_name_spase_table(ws)
    #     get_data(ws, json_arr)
    #     wb.save('БПЛА_parse.xlsx')
    # except Exception as e:
    #     print(f"Ошибка при открытии файла: {e}")



    # # Записываем результаты в xlsx файл
    # wb = openpyxl.Workbook()
    # ws = wb.create_sheet('Drones')
    # worksheet = workbook['имя_листа']

    # проверяем, есть ли на листе строки кроме заголовка





    # 'Максимальная скорость': max_speed,
    # 'Скорость набора','Скорость снижения','Максимальная дальность полета','Максимальная высота полета','Энергопотребление','Масса полезной погрузки','Продолжительность полета','Число винтов'

    #
    # data_dir = os.path.join(os.getcwd(), 'rezDATA')
    # data_file = os.path.join(data_dir, 'BPLA_dji.json')
    # data_file = data_file.replace("\DJIscripts", "")
    # with open(data_file, "w", encoding="utf-8") as outfile:
    #     json.dump(json_arr, outfile, indent=4, ensure_ascii=False)
    #
    # print("Code Complite")

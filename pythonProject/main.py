def write_in_file(categoria,funcNameSt,funcGetData,json_arr):
    try:
        workbook = openpyxl.load_workbook('БПЛА_parse.xlsx')
        print("Файл обрабатывается")
        worksheet = workbook.active
        try:
            worksheet = workbook[categoria]
        except Exception as e:
            worksheet = workbook.create_sheet(categoria)
        if worksheet.max_row == 1:
            funcNameSt(worksheet)
            funcGetData(worksheet, json_arr)
        else:
            funcGetData(worksheet, json_arr)
        # создать список строк, начиная со второй строки (исключая заголовок)
        rows = list(worksheet.iter_rows(min_row=2, values_only=True))
        # создать список уникальных строк
        unique_rows = list(set(map(tuple, rows)))
        # очистить содержимое листа
        worksheet.delete_rows(2, worksheet.max_row)
        # заполнить лист уникальными строками
        for row in unique_rows:
            worksheet.append(row)
        workbook.save('БПЛА_parse.xlsx')
    except FileNotFoundError:
        print("Файл не найден, создадим файл: \'БПЛА_parse.xlsx\'")
        wb = openpyxl.Workbook()
        ws = wb.create_sheet(categoria)
        funcNameSt(ws)
        funcGetData(ws, json_arr)
        wb.save('БПЛА_parse.xlsx')
    except Exception as e:
        print(f"Ошибка при открытии файла: {e}")

import openpyxl

import sys

# проверка на возможность записи в файл
try:
    workbook = openpyxl.load_workbook('БПЛА_parse.xlsx')
    workbook.save('БПЛА_parse.xlsx')
except FileNotFoundError:
    print()
except Exception as e:
    print(f"Ошибка при открытии файла(закройте файл, если он открыт): {e}")
    sys.exit()

from rezDATA import datCateg

import DJIscripts.SITEdji12 as bplaDJI

import GEOBOXscripts.SITEgeobox1 as batteryGEOBOX
import GEOBOXscripts.SITEgeobox6 as lidarGEOBOX
import GEOBOXscripts.SITEgeobox9 as mod_sp_svGEOBOX
import GEOBOXscripts.SITEgeobox11 as teploviGEOBOX
import GEOBOXscripts.SITEgeobox14 as pol_nagrGEOBOX
import GEOBOXscripts.SITEgeobox15 as control_panelGEOBOX

import  AEROMOTUSscript.SITEaeromotus1 as batteryAEROM
import AEROMOTUSscript.SITEaeromotus12 as bplaAEROM
import  AEROMOTUSscript.SITEaeromotus14 as pol_nagrAEROM

bpla_arr = bplaDJI.find_bpla_in_dji()
write_in_file('БПЛА коптерного типа', datCateg.get_name_bpla_table, datCateg.get_data_bpla, bpla_arr)



battery_arr = batteryGEOBOX.fnd_battery_geobox()
write_in_file('Батарея (аккумулятор)', datCateg.get_name_battery_table, datCateg.get_data_battery, battery_arr)

lidar_arr = lidarGEOBOX.find_lidar_geobox()
write_in_file('Лидар', datCateg.get_name_lidar_table, datCateg.get_data_lidar, lidar_arr)

mod_sp_sv_arr = mod_sp_svGEOBOX.find_modul_sp_sv()
write_in_file("Модуль спутниковой связи", datCateg.get_name_modul_sp_sv_table, datCateg.get_data_modul_sp_sv, mod_sp_sv_arr)

teplovi_arr = teploviGEOBOX.find_teplovis_geobox()
write_in_file('Тепловизор', datCateg.get_name_teplovis_table, datCateg.get_data_templovis, teplovi_arr)

pol_nar_arr = pol_nagrGEOBOX.find_pol_nagr_geobox()
write_in_file('Полезная нагрузка', datCateg.get_name_pol_nagr_table, datCateg.get_data_pol_nagr, pol_nar_arr)

control_panel_arr = control_panelGEOBOX.find_pult_uprav_geobox()
write_in_file('Пульт управления', datCateg.get_name_contrl_panel_table, datCateg.get_data_contrl_panel, control_panel_arr)



battery_arr = batteryAEROM.find_battery_aeromotus()
write_in_file('Батарея (аккумулятор)', datCateg.get_name_battery_table, datCateg.get_data_battery, battery_arr)

bpla_arr = bplaAEROM.find_bpla_aeromotus()
write_in_file('БПЛА коптерного типа', datCateg.get_name_bpla_table, datCateg.get_data_bpla, bpla_arr)

pol_nar_arr = pol_nagrAEROM.find_pol_nagr_aeromotus()
write_in_file('Полезная нагрузка', datCateg.get_name_pol_nagr_table, datCateg.get_data_pol_nagr, pol_nar_arr)

print("Конец работы скрипта")